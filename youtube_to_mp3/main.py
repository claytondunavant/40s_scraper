import openpyxl, youtube_dl, os, subprocess
from mutagen.mp3 import MP3
from mutagen.id3 import ID3NoHeaderError
from mutagen.id3 import ID3, TIT2, TALB, TPE1, TPE2, COMM, USLT, TCOM, TCON, TDRC

wb_name = 'songs' #workbook name
wb = openpyxl.load_workbook(wb_name + '.xlsx') #load workbook by name
sheet = wb['Sheet1'] #open first sheet 

os.chdir("./downloads/") #move to downloads directory

ydl_opts = { #download specs
'format': 'bestaudio/best',
'postprocessors': [{
    'key': 'FFmpegExtractAudio',
    'preferredcodec': 'mp3',
    'preferredquality': '192'
    }],
}

def tag_file(fname, row): #tag the mp3 files with their name, artist and album
    # create ID3 tag if not present
    try:
        tags = ID3(fname)
    except ID3NoHeaderError:
        print("Adding ID3 header;"),
        tags = ID3()

    tags["TIT2"] = TIT2(encoding=3, text=str(fname))
    tags["TALB"] = TALB(encoding=3, text=str(sheet['B'+str(row)].value))
    tags["TPE1"] = TPE1(encoding=3, text=str(sheet['A'+str(row)].value))

    tags.save(fname)

old_files = os.listdir() #list all files in downloads

last_downloaded_row = int(input('What was the last downloaded row?: ')) #sets what row downloading will start on 

for i in range(last_downloaded_row + 1, sheet.max_row + 1): #from the one above the last downloaded till the last song
    youtube_dl.YoutubeDL(ydl_opts).download([str(sheet['E' + str(i+1)].value)]) #download the video
    file_name = str(sheet['C' + str(i+1)].value) + '.mp3' #set the file name
    new_files = os.listdir() #look at all the files in downloads
    for o in range(len(new_files)): #for all the file in downloads
        if new_files[o] in old_files: #if the file was in there before, do no thing
            t = 0
        else: #if the file was not in there before
            subprocess.run(['mv', str(new_files[o]), file_name]) #rename the new downloaded file to the desired name
            old_files = os.listdir() #update old files so the new file is in it 
    tag_file(file_name, i+1) #tag the renamed downloaded file 
    print( str( ((i+1) / sheet.max_row) * 100) + '%' + " " + str(i+1) + '/' + str(sheet.max_row)) #print status of how many files are left
